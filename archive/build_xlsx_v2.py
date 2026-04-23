#!/usr/bin/env python3
"""
build_xlsx.py  v3
Reads the latest fundamentals TSV and writes a formatted .xlsx workbook.

Sheets:
  1. Screener     — all protocols, key valuation metrics, heat-map conditional formatting
  2. Revenue      — fee → protocol revenue → earnings → holder revenue waterfall
  3. Multiples    — P/S, P/E, eps_token, earnings_yield, holder_yield comparison
  4. Raw          — unformatted TSV mirror for reference

Usage:
  python3 build_xlsx.py                  # uses latest TSV in output/
  python3 build_xlsx.py output/260331_fundamentals.tsv   # explicit file
"""

import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

ROOT       = Path(__file__).parent
OUTPUT_DIR = ROOT / "output"

# ---------------------------------------------------------------------------
# Colours (hex, no #)
# ---------------------------------------------------------------------------

C_HEADER_BG   = "1A1915"   # near-black
C_HEADER_FG   = "FFFFFF"
C_SECTION_BG  = "F0EDE8"   # light warm grey
C_BORDER      = "D4D0C8"
C_LOW_DQ_BG   = "FDF3E7"   # amber tint for low data-quality rows
C_LOW_DQ_FG   = "7A3F00"
C_GREEN_HI    = "2D5A3D"   # dark green — good metric
C_GREEN_MID   = "EDF5F0"
C_RED_HI      = "8B2500"   # dark red — bad metric
C_RED_MID     = "FDF0EC"
C_NEUTRAL     = "F7F6F2"   # background
C_NONE        = "BBBBBB"   # colour for "—" / missing values
C_TITLE_FG    = "1A1915"

THIN = Side(style="thin", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOTTOM = Border(bottom=Side(style="medium", color=C_BORDER))

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_latest_tsv(explicit=None):
    if explicit:
        p = Path(explicit)
        if not p.exists():
            sys.exit(f"File not found: {explicit}")
        return p
    candidates = sorted(OUTPUT_DIR.glob("??????_fundamentals.tsv"), reverse=True)
    if not candidates:
        sys.exit("No fundamentals TSV found in output/. Run fetch_fundamentals.py first.")
    return candidates[0]


def load_tsv(path):
    with open(path, newline="") as f:
        return list(csv.DictReader(f, delimiter="\t"))


def _f(row, col):
    """Parse float from TSV row; None if empty or non-numeric."""
    v = row.get(col, "")
    try:
        return float(v) if v.strip() != "" else None
    except (ValueError, AttributeError):
        return None


def _fmt_m(val):
    """Format dollar value as $XB / $XM / $XK string, or '—'."""
    if val is None:
        return "—"
    if abs(val) >= 1e9:
        return f"${val/1e9:.2f}B"
    if abs(val) >= 1e6:
        return f"${val/1e6:.2f}M"
    if abs(val) >= 1e3:
        return f"${val/1e3:.1f}K"
    return f"${val:,.0f}"


def _pct(val, decimals=2):
    return f"{val:.{decimals}f}%" if val is not None else "—"


def _mult(val, decimals=1):
    return f"{val:.{decimals}f}x" if val is not None else "—"


def header_fill():
    return PatternFill("solid", fgColor=C_HEADER_BG)


def section_fill():
    return PatternFill("solid", fgColor=C_SECTION_BG)


def low_dq_fill():
    return PatternFill("solid", fgColor=C_LOW_DQ_BG)


def write_cell(ws, row, col, value, bold=False, italic=False,
               align="left", number_format=None, fill=None,
               font_color="000000", font_size=10, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(
        bold=bold, italic=italic, color=font_color, size=font_size,
        name="DM Sans" if not bold else "DM Sans",
    )
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if number_format:
        cell.number_format = number_format
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    return cell


def write_header_row(ws, row_num, labels, col_widths=None):
    """Write a dark header row. labels is list of (col_offset, text, width?)."""
    hf = header_fill()
    for i, label in enumerate(labels):
        c = write_cell(
            ws, row_num, i + 1, label,
            bold=True, fill=hf, font_color=C_HEADER_FG,
            align="center", font_size=9,
        )
        if col_widths and i < len(col_widths):
            ws.column_dimensions[get_column_letter(i + 1)].width = col_widths[i]
    ws.row_dimensions[row_num].height = 20


def write_title(ws, title, subtitle, date_str):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = title
    c.font = Font(name="DM Serif Display", size=16, bold=False, color=C_TITLE_FG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = subtitle + f"  ·  {date_str}"
    c2.font = Font(name="DM Sans", size=9, color="6B6860", italic=True)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 16

    ws.row_dimensions[3].height = 6   # spacer


def freeze_and_zoom(ws, freeze_cell="B5"):
    ws.freeze_panes = freeze_cell
    ws.sheet_view.zoomScale = 100


# ---------------------------------------------------------------------------
# Sheet 1: Screener
# ---------------------------------------------------------------------------

SCREENER_COLS = [
    ("Protocol",       22),
    ("Category",       14),
    ("Fees 30d",       12),
    ("Rev 30d",        12),
    ("MCap",           12),
    ("FDV",            12),
    ("P/S",             8),
    ("P/E",             8),
    ("EPS / Token",    11),
    ("Earn Yld %",     10),
    ("Hld Yld %",      10),
    ("DQ",              5),
]


def build_screener(wb, rows, date_str):
    ws = wb.create_sheet("Screener")
    ws.tab_color = "1A1915"

    write_title(ws,
        "DeFi Fundamental Valuation Screener",
        "Sorted by Earnings Yield (desc) · P/E and Earnings Yield n/a while DL earnings endpoint down",
        date_str)

    labels = [c[0] for c in SCREENER_COLS]
    widths = [c[1] for c in SCREENER_COLS]
    write_header_row(ws, 4, labels, widths)

    data_start = 5
    for i, row in enumerate(rows):
        r = data_start + i
        is_low = row.get("data_quality", "ok") == "low"
        bg = low_dq_fill() if is_low else None
        fg = C_LOW_DQ_FG if is_low else "1A1915"

        def wc(col, value, align="right", nf=None):
            write_cell(ws, r, col, value, fill=bg, font_color=fg,
                       align=align, number_format=nf, font_size=9)

        ps  = _f(row, "ps_ratio")
        pe  = _f(row, "pe_ratio")
        eps = _f(row, "eps_token")
        ey  = _f(row, "earnings_yield")
        hy  = _f(row, "holder_yield")

        wc(1,  row.get("protocol", ""),   align="left")
        wc(2,  row.get("category", ""),   align="left")
        wc(3,  _fmt_m(_f(row, "fees_30d")))
        wc(4,  _fmt_m(_f(row, "revenue_30d")))
        wc(5,  _fmt_m(_f(row, "mcap")))
        wc(6,  _fmt_m(_f(row, "fdv")))
        wc(7,  _mult(ps))
        wc(8,  _mult(pe))
        wc(9,  f"${eps:.4f}" if eps is not None else "—")
        wc(10, _pct(ey))
        wc(11, _pct(hy))
        wc(12, "⚠" if is_low else "", align="center")

        ws.row_dimensions[r].height = 16

    # Zebra striping on non-low-DQ rows
    for i, row in enumerate(rows):
        r = data_start + i
        if row.get("data_quality", "ok") != "low" and i % 2 == 1:
            for col in range(1, len(SCREENER_COLS) + 1):
                cell = ws.cell(row=r, column=col)
                if cell.fill.patternType is None or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = PatternFill("solid", fgColor="F7F6F2")

    # Conditional formatting: P/S (col G = 7) — lower is better (greener)
    last_row = data_start + len(rows) - 1
    ps_range = f"G{data_start}:G{last_row}"
    ws.conditional_formatting.add(ps_range, ColorScaleRule(
        start_type="min", start_color="2D5A3D",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="max", end_color="8B2500",
    ))

    # Earnings yield (col J = 10) — higher is better (greener)
    ey_range = f"J{data_start}:J{last_row}"
    ws.conditional_formatting.add(ey_range, ColorScaleRule(
        start_type="min", start_color="8B2500",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="max", end_color="2D5A3D",
    ))

    # Holder yield (col K = 11) — higher is better (greener)
    hy_range = f"K{data_start}:K{last_row}"
    ws.conditional_formatting.add(hy_range, ColorScaleRule(
        start_type="min", start_color="8B2500",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="max", end_color="2D5A3D",
    ))

    freeze_and_zoom(ws, "C5")


# ---------------------------------------------------------------------------
# Sheet 2: Revenue Waterfall
# ---------------------------------------------------------------------------

REVENUE_COLS = [
    ("Protocol",         22),
    ("Category",         14),
    ("Fees 30d",         12),
    ("→ Protocol Rev",   13),
    ("→ Earnings",       12),
    ("→ Holder Rev",     12),
    ("Ann Fees",         12),
    ("Ann Rev",          12),
    ("Ann Earnings",     13),
    ("Ann Holder Rev",   13),
    ("Take Rate",        10),   # rev/fees — informational only, not a stored metric
    ("Ret Rate",         10),   # holder_rev/revenue — % returned to holders
]


def build_revenue(wb, rows, date_str):
    ws = wb.create_sheet("Revenue")
    ws.tab_color = "2D5A3D"

    write_title(ws,
        "Revenue Waterfall",
        "Fees → Protocol Revenue → Earnings → Holder Revenue · All figures USD",
        date_str)

    labels = [c[0] for c in REVENUE_COLS]
    widths = [c[1] for c in REVENUE_COLS]
    write_header_row(ws, 4, labels, widths)

    data_start = 5
    for i, row in enumerate(rows):
        r = data_start + i
        is_low = row.get("data_quality", "ok") == "low"
        bg  = low_dq_fill() if is_low else None
        fg  = C_LOW_DQ_FG if is_low else "1A1915"

        def wc(col, value, align="right"):
            write_cell(ws, r, col, value, fill=bg, font_color=fg,
                       align=align, font_size=9)

        fees   = _f(row, "fees_30d")
        rev    = _f(row, "revenue_30d")
        earn   = _f(row, "earnings_30d")
        hld    = _f(row, "holders_rev_30d")
        a_fees = _f(row, "ann_fees")
        a_rev  = _f(row, "ann_revenue")
        a_earn = _f(row, "ann_earnings")
        a_hld  = _f(row, "ann_holders_rev")

        # Take rate: rev/fees (informational display only — not a stored metric)
        take = (rev / fees * 100) if (fees and rev is not None and fees != 0) else None
        # Retention rate: holder_rev/protocol_rev
        ret  = (hld / rev * 100) if (rev and hld is not None and rev != 0) else None

        wc(1,  row.get("protocol", ""), align="left")
        wc(2,  row.get("category", ""), align="left")
        wc(3,  _fmt_m(fees))
        wc(4,  _fmt_m(rev))
        wc(5,  _fmt_m(earn))
        wc(6,  _fmt_m(hld))
        wc(7,  _fmt_m(a_fees))
        wc(8,  _fmt_m(a_rev))
        wc(9,  _fmt_m(a_earn))
        wc(10, _fmt_m(a_hld))
        wc(11, _pct(take, 1))
        wc(12, _pct(ret,  1))

        ws.row_dimensions[r].height = 16

    # Zebra
    for i, row in enumerate(rows):
        r = data_start + i
        if row.get("data_quality", "ok") != "low" and i % 2 == 1:
            for col in range(1, len(REVENUE_COLS) + 1):
                cell = ws.cell(row=r, column=col)
                if cell.fill.patternType is None or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = PatternFill("solid", fgColor="F7F6F2")

    freeze_and_zoom(ws, "C5")


# ---------------------------------------------------------------------------
# Sheet 3: Multiples
# ---------------------------------------------------------------------------

MULTIPLES_COLS = [
    ("Protocol",       22),
    ("Category",       14),
    ("MCap",           12),
    ("Ann Revenue",    13),
    ("P/S",             8),
    ("Ann Earnings",   13),
    ("P/E",             8),
    ("EPS / Token",    11),
    ("Earn Yld %",     10),
    ("Hld Yld %",      10),
    ("DQ",              5),
]


def build_multiples(wb, rows, date_str):
    ws = wb.create_sheet("Multiples")
    ws.tab_color = "1E3A5F"

    write_title(ws,
        "Valuation Multiples",
        "P/S uses Protocol Revenue · P/E, EPS, Earnings Yield n/a while DL earnings endpoint unavailable",
        date_str)

    labels = [c[0] for c in MULTIPLES_COLS]
    widths = [c[1] for c in MULTIPLES_COLS]
    write_header_row(ws, 4, labels, widths)

    data_start = 5
    for i, row in enumerate(rows):
        r = data_start + i
        is_low = row.get("data_quality", "ok") == "low"
        bg  = low_dq_fill() if is_low else None
        fg  = C_LOW_DQ_FG if is_low else "1A1915"

        def wc(col, value, align="right"):
            write_cell(ws, r, col, value, fill=bg, font_color=fg,
                       align=align, font_size=9)

        ps  = _f(row, "ps_ratio")
        pe  = _f(row, "pe_ratio")
        eps = _f(row, "eps_token")
        ey  = _f(row, "earnings_yield")
        hy  = _f(row, "holder_yield")

        wc(1,  row.get("protocol", ""),   align="left")
        wc(2,  row.get("category", ""),   align="left")
        wc(3,  _fmt_m(_f(row, "mcap")))
        wc(4,  _fmt_m(_f(row, "ann_revenue")))
        wc(5,  _mult(ps))
        wc(6,  _fmt_m(_f(row, "ann_earnings")))
        wc(7,  _mult(pe))
        wc(8,  f"${eps:.4f}" if eps is not None else "—")
        wc(9,  _pct(ey))
        wc(10, _pct(hy))
        wc(11, "⚠" if is_low else "", align="center")

        ws.row_dimensions[r].height = 16

    last_row = data_start + len(rows) - 1

    # P/S heat map (col E = 5)
    ws.conditional_formatting.add(f"E{data_start}:E{last_row}", ColorScaleRule(
        start_type="min", start_color="2D5A3D",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="max", end_color="8B2500",
    ))
    # Earnings yield (col I = 9)
    ws.conditional_formatting.add(f"I{data_start}:I{last_row}", ColorScaleRule(
        start_type="min", start_color="8B2500",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="max", end_color="2D5A3D",
    ))
    # Holder yield (col J = 10)
    ws.conditional_formatting.add(f"J{data_start}:J{last_row}", ColorScaleRule(
        start_type="min", start_color="8B2500",
        mid_type="percentile", mid_value=50, mid_color="FFFFFF",
        end_type="max", end_color="2D5A3D",
    ))

    # Zebra
    for i, row in enumerate(rows):
        r = data_start + i
        if row.get("data_quality", "ok") != "low" and i % 2 == 1:
            for col in range(1, len(MULTIPLES_COLS) + 1):
                cell = ws.cell(row=r, column=col)
                if cell.fill.patternType is None or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = PatternFill("solid", fgColor="F7F6F2")

    freeze_and_zoom(ws, "C5")


# ---------------------------------------------------------------------------
# Sheet 4: Raw
# ---------------------------------------------------------------------------

def build_raw(wb, rows, date_str):
    ws = wb.create_sheet("Raw")
    ws.tab_color = "6B6860"

    write_title(ws, "Raw Data", "Unformatted TSV mirror — all columns, original values", date_str)

    if not rows:
        return

    # Write header from TSV column names
    cols = list(rows[0].keys())
    hf = header_fill()
    for j, col in enumerate(cols):
        c = ws.cell(row=4, column=j + 1, value=col)
        c.font = Font(bold=True, color=C_HEADER_FG, size=8, name="DM Mono")
        c.fill = hf
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(j + 1)].width = max(10, len(col) + 2)
    ws.row_dimensions[4].height = 18

    for i, row in enumerate(rows):
        r = 5 + i
        for j, col in enumerate(cols):
            val = row.get(col, "")
            # Try to cast numeric strings back to float for proper Excel handling
            cell = ws.cell(row=r, column=j + 1)
            try:
                if val != "":
                    cell.value = float(val)
                    cell.number_format = "0.000000"
                else:
                    cell.value = None
            except (ValueError, TypeError):
                cell.value = val
            cell.font = Font(size=8, name="DM Mono")
            cell.alignment = Alignment(horizontal="right" if j > 3 else "left")
        ws.row_dimensions[r].height = 14

    ws.freeze_panes = "A5"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    explicit = sys.argv[1] if len(sys.argv) > 1 else None
    tsv_path = find_latest_tsv(explicit)
    rows = load_tsv(tsv_path)

    if not rows:
        sys.exit(f"No data rows in {tsv_path}")

    date_str = rows[0].get("date", tsv_path.stem[:6])
    print(f"Building workbook from {tsv_path.name}  ({len(rows)} protocols)…")

    # Sort: earnings_yield desc (None last), then revenue_30d desc
    rows.sort(key=lambda r: (
        -(_f(r, "earnings_yield") or 0) if _f(r, "earnings_yield") is not None else float("inf"),
        -(_f(r, "revenue_30d") or 0),
    ))

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    build_screener(wb, rows, date_str)
    build_revenue(wb, rows, date_str)
    build_multiples(wb, rows, date_str)
    build_raw(wb, rows, date_str)

    out_path = OUTPUT_DIR / f"{tsv_path.stem.split('_')[0]}_fundamentals.xlsx"
    wb.save(out_path)
    print(f"Saved → {out_path.name}")


if __name__ == "__main__":
    main()
