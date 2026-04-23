#!/usr/bin/env python3
"""
build_xlsx.py  v3
Reads latest DeFi fundamentals TSV + equities TSV, writes formatted .xlsx.

Sheets:
  1. DeFi Protocols    — 25 DeFi protocols, key valuation metrics
  2. Tech Equities     — 25 tech/fintech equities
  3. DeFi vs Equities  — side-by-side comparison (Excel formulas for derived cells)
  4. By Sector         — DeFi sector benchmarks
  5. Data Quality      — completeness audit for all rows

Usage:
  python3 build_xlsx.py                           # auto-detect latest TSVs
  python3 build_xlsx.py --defi output/260401_fundamentals.tsv
  python3 build_xlsx.py --equities output/260401_equities.tsv
"""

import csv
import statistics
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

ROOT       = Path(__file__).parent
OUTPUT_DIR = ROOT / "output"

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------

C_HEADER_BG    = "1A1915"
C_HEADER_FG    = "FFFFFF"
C_ALT_ROW      = "F7F6F2"
C_GREEN_FILL   = "C6EFCE"
C_GREEN_FONT   = "276221"
C_RED_FILL     = "FFC7CE"
C_RED_FONT     = "9C0006"
C_ORANGE_FILL  = "FFEB9C"
C_ORANGE_FONT  = "7A3F00"
C_TAB_DEFI     = "2D5A3D"
C_TAB_EQUITY   = "1E3A5F"
C_TAB_COMPARE  = "4A1942"
C_TAB_SECTOR   = "7A3F00"
C_TAB_QUALITY  = "444441"
C_BORDER       = "D4D0C8"

# ---------------------------------------------------------------------------
# Number formats
# ---------------------------------------------------------------------------

FMT_DOLLAR     = '$#,##0'
FMT_MULTIPLE   = '0.0"x"'
FMT_PERCENT    = '0.0%'
FMT_PERCENT_2  = '0.00%'
FMT_EPS_TOKEN  = '$0.0000'
FMT_EPS_EQ     = '$0.00'
FMT_MOMENTUM   = '"▲"0.0%;"▼"0.0%;"-"'

# ---------------------------------------------------------------------------
# Shared styles
# ---------------------------------------------------------------------------

THIN   = Side(style="thin",   color=C_BORDER)
MEDIUM = Side(style="medium", color=C_BORDER)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def _header_fill():
    return PatternFill("solid", fgColor=C_HEADER_BG)

def _alt_fill():
    return PatternFill("solid", fgColor=C_ALT_ROW)

def _header_font(size=11):
    return Font(name="Arial", bold=True, color=C_HEADER_FG, size=size)

def _body_font(bold=False):
    return Font(name="Arial", bold=bold, size=10)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")

def _left():
    return Alignment(horizontal="left", vertical="center")

# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def find_latest_tsv(pattern):
    candidates = sorted(OUTPUT_DIR.glob(pattern), reverse=True)
    return candidates[0] if candidates else None

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _f(row, col):
    """Parse float from TSV row dict. Returns None if empty or non-numeric."""
    v = row.get(col, "")
    if not v or str(v).strip() == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def load_tsv(path):
    if not path or not path.exists():
        return []
    with open(path, newline="") as f:
        return list(csv.DictReader(f, delimiter="\t"))


def load_defi(path):
    rows = load_tsv(path)
    # Already sorted by earnings_yield desc from fetch_fundamentals.py
    return rows


def load_equities(path):
    rows = load_tsv(path)
    # Sort by earnings_yield desc, None last
    def ey_sort(r):
        v = _f(r, "earnings_yield")
        return (0, -(v or 0)) if v is not None else (1, 0)
    rows.sort(key=ey_sort)
    return rows

# ---------------------------------------------------------------------------
# Cell-writing helpers
# ---------------------------------------------------------------------------

DASH = "—"

def cell_num(val):
    """Return numeric value or DASH sentinel for None."""
    return val if val is not None else DASH

def pct_to_dec(val):
    """Convert TSV percent value (e.g. 41.68) to Excel decimal (0.4168). None-safe."""
    return val / 100.0 if val is not None else None


def apply_header_row(ws, row_num, headers, widths=None):
    """Write a header row with dark fill / white bold text."""
    fill = _header_fill()
    font = _header_font()
    for col_idx, text in enumerate(headers, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=text)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = _center()
        cell.border = BORDER
        if widths and col_idx - 1 < len(widths):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths[col_idx - 1]


def write_data_row(ws, row_num, values, formats=None, is_alt=False):
    """
    Write a data row. values is a list of cell values (None → DASH string).
    formats: list of format strings per column (None = default).
    """
    fill = _alt_fill() if is_alt else None
    for col_idx, val in enumerate(values, start=1):
        if val is None:
            display = DASH
        else:
            display = val
        cell = ws.cell(row=row_num, column=col_idx, value=display)
        cell.font      = _body_font()
        cell.alignment = _right() if isinstance(display, (int, float)) else _left()
        cell.border    = BORDER
        if fill:
            cell.fill = fill
        if formats and col_idx - 1 < len(formats) and formats[col_idx - 1]:
            if isinstance(display, (int, float)):
                cell.number_format = formats[col_idx - 1]
    return


def apply_cf_tercile(ws, col_letter, data_start, data_end, values):
    """
    Apply green (top third) / red (bottom third) conditional formatting
    to a column of percentage-as-decimal values.
    """
    nums = [v for v in values if v is not None]
    if len(nums) < 3:
        return
    nums_sorted = sorted(nums)
    n = len(nums_sorted)
    low_thresh  = nums_sorted[n // 3]
    high_thresh = nums_sorted[(2 * n) // 3]
    cf_range    = f"{col_letter}{data_start}:{col_letter}{data_end}"

    green_rule = CellIsRule(
        operator="greaterThanOrEqual",
        formula=[str(high_thresh)],
        fill=PatternFill("solid", fgColor=C_GREEN_FILL),
        font=Font(name="Arial", color=C_GREEN_FONT, size=10),
    )
    red_rule = CellIsRule(
        operator="lessThanOrEqual",
        formula=[str(low_thresh)],
        fill=PatternFill("solid", fgColor=C_RED_FILL),
        font=Font(name="Arial", color=C_RED_FONT, size=10),
    )
    ws.conditional_formatting.add(cf_range, green_rule)
    ws.conditional_formatting.add(cf_range, red_rule)


def configure_sheet(ws, tab_color, freeze="A2"):
    ws.sheet_properties.tabColor   = tab_color
    ws.sheet_view.showGridLines    = False
    if freeze:
        ws.freeze_panes = freeze

# ---------------------------------------------------------------------------
# Sheet 1 — DeFi Protocols
# ---------------------------------------------------------------------------

def write_defi_sheet(wb, defi_rows):
    ws = wb.create_sheet("DeFi Protocols")
    configure_sheet(ws, C_TAB_DEFI)

    headers = [
        "Protocol", "Category", "Market Cap", "Ann Revenue", "Ann Earnings",
        "Ann Holder Rev", "P/S", "P/E", "Earnings/Token",
        "Earnings Yield", "Holder Yield",
        "Rev Momentum", "Supply Inflation", "Real Yield",
        "Data Quality",
    ]
    widths = [18, 16, 13, 13, 13, 13, 13, 13, 13, 13, 13, 13, 14, 11, 12]
    formats = [
        None, None,
        FMT_DOLLAR, FMT_DOLLAR, FMT_DOLLAR, FMT_DOLLAR,
        FMT_MULTIPLE, FMT_MULTIPLE, FMT_EPS_TOKEN,
        FMT_PERCENT_2, FMT_PERCENT_2,
        FMT_MOMENTUM, FMT_PERCENT, FMT_PERCENT,
        None,
    ]

    apply_header_row(ws, 1, headers, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 18

    ey_vals, mom_vals, si_vals, ry_vals = [], [], [], []
    for row_idx, r in enumerate(defi_rows, start=2):
        ey_tsv  = _f(r, "earnings_yield")
        hy_tsv  = _f(r, "holder_yield")
        mom_raw = _f(r, "revenue_momentum")   # already decimal (0.12 = 12%)
        si_raw  = _f(r, "supply_inflation")   # percent (5.0 = 5%)
        ry_raw  = _f(r, "real_yield")         # percent

        ey_dec  = pct_to_dec(ey_tsv)
        hy_dec  = pct_to_dec(hy_tsv)
        si_dec  = pct_to_dec(si_raw)
        ry_dec  = pct_to_dec(ry_raw)

        if ey_dec  is not None: ey_vals.append(ey_dec)
        if mom_raw is not None: mom_vals.append(mom_raw)
        if si_dec  is not None: si_vals.append(si_dec)
        if ry_dec  is not None: ry_vals.append(ry_dec)

        values = [
            r.get("protocol"),
            r.get("category"),
            _f(r, "mcap"),
            _f(r, "ann_revenue"),
            _f(r, "ann_earnings"),
            _f(r, "ann_holders_rev"),
            _f(r, "ps_ratio"),
            _f(r, "pe_ratio"),
            _f(r, "eps_token"),
            ey_dec,
            hy_dec,
            mom_raw,   # decimal — FMT_MOMENTUM displays as ▲/▼ %
            si_dec,
            ry_dec,
            r.get("data_quality"),
        ]
        write_data_row(ws, row_idx, values, formats, is_alt=(row_idx % 2 == 0))
        # Left-align text columns (Protocol=1, Category=2, Data Quality=15)
        for c in [1, 2, 15]:
            ws.cell(row=row_idx, column=c).alignment = _left()

    last_row = 1 + len(defi_rows)

    # Conditional formatting on Earnings Yield (col J = 10)
    apply_cf_tercile(ws, "J", 2, last_row, ey_vals)

    # Rev Momentum (col L = 12): green >5%, red <-5%
    mom_range = f"L2:L{last_row}"
    ws.conditional_formatting.add(mom_range, CellIsRule(
        operator="greaterThan", formula=["0.05"],
        fill=PatternFill("solid", fgColor=C_GREEN_FILL),
        font=Font(name="Arial", color=C_GREEN_FONT, size=10),
    ))
    ws.conditional_formatting.add(mom_range, CellIsRule(
        operator="lessThan", formula=["-0.05"],
        fill=PatternFill("solid", fgColor=C_RED_FILL),
        font=Font(name="Arial", color=C_RED_FONT, size=10),
    ))

    # Supply Inflation (col M = 13): amber if >10%
    si_range = f"M2:M{last_row}"
    ws.conditional_formatting.add(si_range, CellIsRule(
        operator="greaterThan", formula=["0.10"],
        fill=PatternFill("solid", fgColor=C_ORANGE_FILL),
        font=Font(name="Arial", color=C_ORANGE_FONT, size=10),
    ))

    # Real Yield (col N = 14): green if >0, red if <0
    ry_range = f"N2:N{last_row}"
    ws.conditional_formatting.add(ry_range, CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", fgColor=C_GREEN_FILL),
        font=Font(name="Arial", color=C_GREEN_FONT, size=10),
    ))
    ws.conditional_formatting.add(ry_range, CellIsRule(
        operator="lessThan", formula=["0"],
        fill=PatternFill("solid", fgColor=C_RED_FILL),
        font=Font(name="Arial", color=C_RED_FONT, size=10),
    ))

    return ws, last_row


# ---------------------------------------------------------------------------
# Sheet 2 — Tech Equities
# ---------------------------------------------------------------------------

def write_equities_sheet(wb, eq_rows):
    ws = wb.create_sheet("Tech Equities")
    configure_sheet(ws, C_TAB_EQUITY)

    headers = [
        "Ticker", "Company", "Sector", "Market Cap", "Revenue (TTM)",
        "Net Income (TTM)", "EPS", "P/E", "P/S",
        "Earnings Yield", "Dividend Yield", "Gross Margin", "Operating Margin",
        "Rev Momentum", "Supply Inflation", "Real Yield",
    ]
    widths = [8, 24, 16, 13, 13, 13, 10, 10, 10, 13, 13, 13, 13, 13, 14, 11]
    formats = [
        None, None, None,
        FMT_DOLLAR, FMT_DOLLAR, FMT_DOLLAR,
        FMT_EPS_EQ,
        FMT_MULTIPLE, FMT_MULTIPLE,
        FMT_PERCENT_2, FMT_PERCENT_2,
        FMT_PERCENT, FMT_PERCENT,
        FMT_MOMENTUM, FMT_PERCENT, FMT_PERCENT,
    ]

    apply_header_row(ws, 1, headers, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 18

    ey_vals = []
    for row_idx, r in enumerate(eq_rows, start=2):
        ey_tsv  = _f(r, "earnings_yield")
        div_tsv = _f(r, "dividend_yield_pct")
        gm      = _f(r, "gross_margins")
        om      = _f(r, "operating_margins")
        mom_raw = _f(r, "revenue_momentum")
        si_raw  = _f(r, "supply_inflation")
        ry_raw  = _f(r, "real_yield")

        ey_dec  = pct_to_dec(ey_tsv)
        div_dec = pct_to_dec(div_tsv)
        si_dec  = pct_to_dec(si_raw)
        ry_dec  = pct_to_dec(ry_raw)

        if ey_dec is not None:
            ey_vals.append(ey_dec)

        values = [
            r.get("ticker"),
            r.get("company"),
            r.get("sector"),
            _f(r, "market_cap"),
            _f(r, "total_revenue_ttm"),
            _f(r, "net_income_ttm"),
            _f(r, "trailing_eps"),
            _f(r, "pe_ratio"),
            _f(r, "ps_ratio"),
            ey_dec,
            div_dec,
            gm,
            om,
            mom_raw,
            si_dec,
            ry_dec,
        ]
        write_data_row(ws, row_idx, values, formats, is_alt=(row_idx % 2 == 0))
        for c in [1, 2, 3]:
            ws.cell(row=row_idx, column=c).alignment = _left()

    last_row = 1 + len(eq_rows)
    apply_cf_tercile(ws, "J", 2, last_row, ey_vals)

    # Rev Momentum (col N = 14): green >5%, red <-5%
    mom_range = f"N2:N{last_row}"
    ws.conditional_formatting.add(mom_range, CellIsRule(
        operator="greaterThan", formula=["0.05"],
        fill=PatternFill("solid", fgColor=C_GREEN_FILL),
        font=Font(name="Arial", color=C_GREEN_FONT, size=10),
    ))
    ws.conditional_formatting.add(mom_range, CellIsRule(
        operator="lessThan", formula=["-0.05"],
        fill=PatternFill("solid", fgColor=C_RED_FILL),
        font=Font(name="Arial", color=C_RED_FONT, size=10),
    ))

    # Supply Inflation (col O = 15): amber if >10%
    si_range = f"O2:O{last_row}"
    ws.conditional_formatting.add(si_range, CellIsRule(
        operator="greaterThan", formula=["0.10"],
        fill=PatternFill("solid", fgColor=C_ORANGE_FILL),
        font=Font(name="Arial", color=C_ORANGE_FONT, size=10),
    ))

    # Real Yield (col P = 16): green if >0, red if <0
    ry_range = f"P2:P{last_row}"
    ws.conditional_formatting.add(ry_range, CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", fgColor=C_GREEN_FILL),
        font=Font(name="Arial", color=C_GREEN_FONT, size=10),
    ))
    ws.conditional_formatting.add(ry_range, CellIsRule(
        operator="lessThan", formula=["0"],
        fill=PatternFill("solid", fgColor=C_RED_FILL),
        font=Font(name="Arial", color=C_RED_FONT, size=10),
    ))

    return ws, last_row


# ---------------------------------------------------------------------------
# Sheet 3 — DeFi vs Equities
# ---------------------------------------------------------------------------

def write_comparison_sheet(wb, defi_last_row, eq_last_row):
    """
    Derived cells (DeFi vs Equity ratio column) use Excel formulas.
    Aggregate stats (median/min/max) also use Excel formulas referencing
    the data sheets, so values update automatically if underlying data changes.
    """
    ws = wb.create_sheet("DeFi vs Equities")
    configure_sheet(ws, C_TAB_COMPARE)

    # Column widths
    ws.column_dimensions["A"].width = 24
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 16

    # Title row
    title_cell = ws.cell(row=1, column=1, value=f"DeFi vs Equities Comparison")
    title_cell.font      = Font(name="Arial", bold=True, size=13, color=C_HEADER_BG)
    title_cell.alignment = _left()
    ws.row_dimensions[1].height = 22

    # ── Main comparison table ──
    tbl_header_row = 3
    tbl_data_start = 4

    cmp_headers = [
        "Metric",
        "DeFi Median", "DeFi Min", "DeFi Max",
        "Equity Median", "Equity Min", "Equity Max",
        "DeFi vs Equity",
    ]
    # Sub-header (row 3)
    fill = _header_fill()
    font = _header_font()
    for col_idx, text in enumerate(cmp_headers, start=1):
        cell = ws.cell(row=tbl_header_row, column=col_idx, value=text)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = _center()
        cell.border = BORDER
    ws.row_dimensions[tbl_header_row].height = 18

    # DeFi data ranges (Sheets 1 and 2 row 2 to last_row)
    d_end = max(defi_last_row, 26)   # at least row 26 for safety
    e_end = max(eq_last_row, 26)

    # Metric definitions:
    # (label, defi_col, equity_col, row_fmt)
    # defi_col / equity_col are the column letters in Sheets 1 and 2
    # row_fmt: "multiple" | "percent" | "percent1" | "eps"
    metrics = [
        ("P/S Ratio",                    "G", "I", "multiple"),
        ("P/E Ratio",                    "H", "H", "multiple"),
        ("Earnings Yield",               "J", "J", "percent"),
        ("Holder / Dividend Yield",      "K", "K", "percent"),
        ("EPS (indexed, USD/unit)",      "I", "G", "eps"),
        ("Revenue Momentum",             "L", "N", "percent1"),
        ("Supply Inflation",             "M", "O", "percent1"),
        ("Real Yield",                   "N", "P", "percent1"),
    ]

    row_fmts = {
        "multiple": FMT_MULTIPLE,
        "percent":  FMT_PERCENT_2,
        "percent1": FMT_PERCENT,
        "eps":      FMT_EPS_EQ,
    }

    for i, (label, d_col, e_col, fmt_key) in enumerate(metrics):
        r = tbl_data_start + i
        fmt = row_fmts[fmt_key]
        is_alt = (i % 2 == 1)
        row_fill = _alt_fill() if is_alt else None

        d_range = f"'DeFi Protocols'!{d_col}2:{d_col}{d_end}"
        e_range = f"'Tech Equities'!{e_col}2:{e_col}{e_end}"

        row_data = [
            # A: label
            (label, None),
            # B: DeFi Median
            (f"=IFERROR(MEDIAN({d_range}),\"—\")", fmt),
            # C: DeFi Min — MIN ignores text "—" cells naturally
            (f"=MIN({d_range})", fmt),
            # D: DeFi Max
            (f"=MAX({d_range})", fmt),
            # E: Equity Median
            (f"=IFERROR(MEDIAN({e_range}),\"—\")", fmt),
            # F: Equity Min
            (f"=MIN({e_range})", fmt),
            # G: Equity Max
            (f"=MAX({e_range})", fmt),
            # H: DeFi vs Equity ratio (formula)
            ('=IFERROR(IF(E{r}=0,"—",B{r}/E{r}),"—")'.replace("{r}", str(r)), FMT_MULTIPLE),
        ]

        for col_idx, (val, num_fmt) in enumerate(row_data, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font      = _body_font()
            cell.border    = BORDER
            cell.alignment = _left() if col_idx == 1 else _center()
            if row_fill:
                cell.fill = row_fill
            if num_fmt and col_idx > 1:
                cell.number_format = num_fmt

        ws.row_dimensions[r].height = 16

    # ── Top 5 sub-table ──
    sub_start_row = tbl_data_start + len(metrics) + 2

    # Sub-table header
    sub_title = ws.cell(row=sub_start_row, column=1, value="Top 5 by Earnings Yield")
    sub_title.font = Font(name="Arial", bold=True, size=11, color=C_HEADER_BG)
    ws.row_dimensions[sub_start_row].height = 18

    sub_header_row = sub_start_row + 1
    sub_headers = [
        "DeFi Protocol", "Category", "DeFi Earn Yld", "DeFi P/E",
        "",
        "Equity Ticker", "Company", "Equity Earn Yld", "Equity P/E",
    ]
    for col_idx, text in enumerate(sub_headers, start=1):
        cell = ws.cell(row=sub_header_row, column=col_idx, value=text)
        if text:
            cell.fill  = _header_fill()
            cell.font  = _header_font()
            cell.alignment = _center()
            cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions["A"].width = 24   # restore widths overridden above
    ws.column_dimensions["E"].width = 2    # spacer column

    # Top 5 rows — formula references to Sheets 1 and 2
    for rank in range(1, 6):
        r = sub_header_row + rank
        src_row = rank + 1  # Sheet 1/2 data starts at row 2
        is_alt  = (rank % 2 == 0)
        row_fill = _alt_fill() if is_alt else None

        sub_row = [
            (f"='DeFi Protocols'!A{src_row}", None),
            (f"='DeFi Protocols'!B{src_row}", None),
            (f"='DeFi Protocols'!J{src_row}", FMT_PERCENT_2),
            (f"='DeFi Protocols'!H{src_row}", FMT_MULTIPLE),
            ("", None),  # spacer
            (f"='Tech Equities'!A{src_row}", None),
            (f"='Tech Equities'!B{src_row}", None),
            (f"='Tech Equities'!J{src_row}", FMT_PERCENT_2),
            (f"='Tech Equities'!H{src_row}", FMT_MULTIPLE),
        ]
        for col_idx, (val, num_fmt) in enumerate(sub_row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=val)
            cell.font      = _body_font()
            cell.alignment = _left() if col_idx in (1, 2, 6, 7) else _center()
            if col_idx != 5:
                cell.border = BORDER
            if row_fill and col_idx != 5:
                cell.fill = row_fill
            if num_fmt:
                cell.number_format = num_fmt

    return ws


# ---------------------------------------------------------------------------
# Sheet 4 — By Sector
# ---------------------------------------------------------------------------

def write_sector_sheet(wb, defi_rows):
    ws = wb.create_sheet("By Sector")
    configure_sheet(ws, C_TAB_SECTOR)

    # Aggregate by category
    from collections import defaultdict
    sectors = defaultdict(list)
    for r in defi_rows:
        sectors[r.get("category", "Unknown")].append(r)

    # Compute stats per sector
    def med(vals):
        nums = [v for v in vals if v is not None]
        return statistics.median(nums) if nums else None

    sector_stats = []
    for cat, rows in sectors.items():
        ps_vals  = [_f(r, "ps_ratio")       for r in rows]
        pe_vals  = [_f(r, "pe_ratio")        for r in rows]
        ey_vals  = [pct_to_dec(_f(r, "earnings_yield")) for r in rows]
        hy_vals  = [pct_to_dec(_f(r, "holder_yield"))   for r in rows]
        rev_vals = [_f(r, "ann_revenue")     for r in rows]
        earn_vals= [_f(r, "ann_earnings")    for r in rows]

        total_rev  = sum(v for v in rev_vals  if v is not None) or None
        total_earn = sum(v for v in earn_vals if v is not None) or None

        sector_stats.append({
            "category":      cat,
            "count":         len(rows),
            "median_ps":     med(ps_vals),
            "median_pe":     med(pe_vals),
            "median_ey":     med(ey_vals),
            "median_hy":     med(hy_vals),
            "total_rev":     total_rev,
            "total_earn":    total_earn,
        })

    # Sort by median earnings yield desc (None last)
    sector_stats.sort(key=lambda s: (
        0 if s["median_ey"] is not None else 1,
        -(s["median_ey"] or 0),
    ))

    headers = [
        "Category", "Protocols",
        "Median P/S", "Median P/E", "Median Earn Yld", "Median Holder Yld",
        "Total Ann Revenue", "Total Ann Earnings",
    ]
    widths = [18, 10, 13, 13, 15, 15, 18, 18]
    formats = [
        None, None,
        FMT_MULTIPLE, FMT_MULTIPLE, FMT_PERCENT_2, FMT_PERCENT_2,
        FMT_DOLLAR, FMT_DOLLAR,
    ]

    apply_header_row(ws, 1, headers, widths)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.row_dimensions[1].height = 18

    for row_idx, s in enumerate(sector_stats, start=2):
        values = [
            s["category"],
            s["count"],
            s["median_ps"],
            s["median_pe"],
            s["median_ey"],
            s["median_hy"],
            s["total_rev"],
            s["total_earn"],
        ]
        write_data_row(ws, row_idx, values, formats, is_alt=(row_idx % 2 == 0))
        ws.cell(row=row_idx, column=1).alignment = _left()

    return ws


# ---------------------------------------------------------------------------
# Sheet 5 — Data Quality
# ---------------------------------------------------------------------------

def write_quality_sheet(wb, defi_rows, eq_rows):
    ws = wb.create_sheet("Data Quality")
    configure_sheet(ws, C_TAB_QUALITY)

    headers = ["Name", "Type", "Data Source", "Missing Count", "Missing Fields", "Notes"]
    widths  = [24, 10, 14, 14, 40, 40]
    apply_header_row(ws, 1, headers, widths)
    ws.row_dimensions[1].height = 18

    DEFI_KEY_FIELDS  = ["mcap", "ann_revenue", "ann_earnings", "ps_ratio", "pe_ratio",
                        "earnings_yield", "holder_yield", "eps_token"]
    EQUITY_KEY_FIELDS = ["market_cap", "total_revenue_ttm", "net_income_ttm",
                         "pe_ratio", "ps_ratio", "earnings_yield", "trailing_eps",
                         "gross_margins", "operating_margins"]

    orange_fill = PatternFill("solid", fgColor=C_ORANGE_FILL)
    orange_font = Font(name="Arial", color=C_ORANGE_FONT, size=10)

    all_rows = (
        [(r, "DeFi",   DEFI_KEY_FIELDS)   for r in defi_rows] +
        [(r, "Equity", EQUITY_KEY_FIELDS) for r in eq_rows]
    )

    for excel_row, (r, row_type, key_fields) in enumerate(all_rows, start=2):
        missing = [f for f in key_fields if not r.get(f) and _f(r, f) is None]
        name    = r.get("protocol") or r.get("ticker") or "unknown"
        source  = r.get("data_source") or r.get("data_quality") or ""
        notes   = r.get("notes", "")
        is_alt  = (excel_row % 2 == 0)

        values = [name, row_type, source, len(missing), ", ".join(missing), notes]
        write_data_row(ws, excel_row, values, None, is_alt=is_alt)
        for c in [1, 2, 3, 5, 6]:
            ws.cell(row=excel_row, column=c).alignment = _left()

        # Flag rows with > 3 missing fields in orange
        if len(missing) > 3:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=excel_row, column=col)
                cell.fill = orange_fill
                cell.font = orange_font

    return ws


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    # ── Parse optional CLI args ──
    defi_path = eq_path = None
    args = sys.argv[1:]
    for i, arg in enumerate(args):
        if arg == "--defi" and i + 1 < len(args):
            defi_path = Path(args[i + 1])
        elif arg == "--equities" and i + 1 < len(args):
            eq_path = Path(args[i + 1])

    if not defi_path:
        defi_path = find_latest_tsv("??????_fundamentals.tsv")
    if not eq_path:
        eq_path = find_latest_tsv("??????_equities.tsv")

    if not defi_path or not defi_path.exists():
        sys.exit("ERROR: No fundamentals TSV found. Run fetch_fundamentals.py first.")

    print(f"DeFi data  : {defi_path.name}")
    if eq_path and eq_path.exists():
        print(f"Equity data: {eq_path.name}")
    else:
        print("Equity data: NOT FOUND — Sheet 2 will be empty")
        eq_path = None

    # Derive output path from defi TSV date prefix
    date_prefix = defi_path.name[:6]
    xlsx_path   = OUTPUT_DIR / f"{date_prefix}_fundamentals.xlsx"

    defi_rows = load_defi(defi_path)
    eq_rows   = load_equities(eq_path) if eq_path else []

    print(f"Loaded {len(defi_rows)} DeFi rows, {len(eq_rows)} equity rows")

    # ── Build workbook ──
    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    print("Building Sheet 1: DeFi Protocols…")
    _, defi_last_row = write_defi_sheet(wb, defi_rows)

    print("Building Sheet 2: Tech Equities…")
    _, eq_last_row = write_equities_sheet(wb, eq_rows)

    print("Building Sheet 3: DeFi vs Equities…")
    write_comparison_sheet(wb, defi_last_row, eq_last_row)

    print("Building Sheet 4: By Sector…")
    write_sector_sheet(wb, defi_rows)

    print("Building Sheet 5: Data Quality…")
    write_quality_sheet(wb, defi_rows, eq_rows)

    # ── Save ──
    wb.save(xlsx_path)
    print(f"\nSaved → {xlsx_path.name}")

    # ── Summary ──
    print("\n── Sheet summary ──")
    for sheet in wb.sheetnames:
        ws  = wb[sheet]
        nrows = ws.max_row - 1  # subtract header
        print(f"  {sheet:<22} {nrows} data rows")

    missing_defi  = sum(1 for r in defi_rows if not _f(r, "pe_ratio"))
    missing_eq    = sum(1 for r in eq_rows   if not _f(r, "pe_ratio"))
    if missing_defi:
        print(f"\n  ⚠ {missing_defi} DeFi protocols missing P/E")
    if missing_eq:
        print(f"  ⚠ {missing_eq} equities missing P/E")

    # ── Momentum & Real Yield summary ──
    print("\n── Revenue Momentum & Real Yield ──")
    print(f"  {'Name':<26} {'Type':<6} {'Rev Momentum':>14} {'Real Yield':>11}")
    print("  " + "─" * 60)

    flagged = []
    all_rows = (
        [(r, "DeFi", r.get("protocol", "?")) for r in defi_rows] +
        [(r, "EQ",   r.get("ticker",   "?")) for r in eq_rows]
    )
    for r, row_type, name in all_rows:
        rm  = _f(r, "revenue_momentum")
        ry  = _f(r, "real_yield")
        ei  = _f(r, "earnings_yield")
        si  = _f(r, "supply_inflation")

        rm_str = f"{rm*100:+.1f}%" if rm is not None else "—"
        ry_str = f"{ry:+.1f}%"     if ry is not None else "—"
        print(f"  {name:<26} {row_type:<6} {rm_str:>14} {ry_str:>11}")

        if ei is not None and ei > 0 and si is not None and si > ei:
            flagged.append((name, row_type, ei, si, ry))

    print()
    if flagged:
        print("  ⚠ Supply inflation exceeds earnings yield (real yield negative despite positive earnings):")
        for name, row_type, ei, si, ry in flagged:
            ry_disp = f"{ry:+.1f}%" if ry is not None else "—"
            print(f"    {name} ({row_type}):  earn_yld={ei:.1f}%  supply_inf={si:.1f}%  real_yield={ry_disp}")
    else:
        print("  ✓ No assets have supply inflation exceeding earnings yield.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
